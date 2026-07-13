from __future__ import annotations

import hashlib
import io
import json
import os
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from fmr.workbook.inspect import inspect_workbook_bytes
from fmr.workbook.write_plan_public import validate_workbook_write_plan_payload

_EXECUTION_CONTROLS = (
    "atomic_output_publish",
    "failed_output_removed",
    "formulas_not_calculated",
    "no_source_overwrite",
    "output_reopened_and_verified",
    "receipt_excludes_cell_values",
    "source_hash_verified",
    "write_plan_validated",
)


@dataclass(frozen=True)
class WorkbookExecutionResult:
    output_bytes: bytes
    receipt: dict[str, Any]


def execute_workbook_write_plan_bytes(
    source_bytes: bytes,
    *,
    filename: str,
    output_filename: str,
    write_plan: dict[str, Any],
) -> WorkbookExecutionResult:
    _require_openpyxl()
    if not isinstance(source_bytes, bytes) or not source_bytes:
        raise ValueError("source workbook bytes must be non-empty")
    if not filename.lower().endswith(".xlsx"):
        raise ValueError("source workbook must use the .xlsx extension")
    if not output_filename.lower().endswith(".xlsx"):
        raise ValueError("output workbook must use the .xlsx extension")

    issues = validate_workbook_write_plan_payload(write_plan)
    if issues:
        raise ValueError("invalid workbook write plan: " + "; ".join(issues))
    if not write_plan.get("ready_for_executor"):
        raise ValueError("workbook write plan is blocked")

    source_sha256 = hashlib.sha256(source_bytes).hexdigest()
    source = write_plan["source"]
    if source_sha256 != source["sha256"]:
        raise ValueError("source workbook hash does not match the write plan")
    if len(source_bytes) != source["size_bytes"]:
        raise ValueError("source workbook size does not match the write plan")

    source_map = inspect_workbook_bytes(source_bytes, filename=filename)
    if source_map.workbook.external_links_detected:
        raise ValueError("source workbook contains external links")
    if source_map.limitations:
        unsafe = [
            item
            for item in source_map.limitations
            if any(token in item.lower() for token in ("macro", "encrypted", "unsupported"))
        ]
        if unsafe:
            raise ValueError("source workbook contains unsupported features: " + "; ".join(unsafe))

    workbook = _load_workbook(source_bytes)
    record_receipts: list[dict[str, Any]] = []
    try:
        for phase in write_plan["phases"]:
            for record in phase["records"]:
                before = _record_state(workbook, record)
                _apply_record(workbook, record)
                after = _record_state(workbook, record)
                record_receipts.append(
                    {
                        "record_id": record["record_id"],
                        "sequence": record["sequence"],
                        "write_kind": record["write_kind"],
                        "sheet_name": record["sheet_name"],
                        "coordinate": record["coordinate"],
                        "status": "applied",
                        "cell_count": after["cell_count"],
                        "before_sha256": _digest(before),
                        "after_sha256": _digest(after),
                    }
                )
        _request_recalculation(workbook)
        output_bytes = _save_workbook(workbook)
    finally:
        workbook.close()

    output_map = inspect_workbook_bytes(output_bytes, filename=output_filename)
    if output_map.workbook.external_links_detected:
        raise ValueError("executor produced an output workbook with external links")

    reopened = _load_workbook(output_bytes)
    try:
        verification = _verify_records(reopened, write_plan)
    finally:
        reopened.close()
    if verification["failed_record_ids"]:
        raise ValueError(
            "output workbook verification failed for records: "
            + ", ".join(verification["failed_record_ids"])
        )

    output_sha256 = hashlib.sha256(output_bytes).hexdigest()
    provisional = {
        "contract_version": "workbook-execution-receipt.v1",
        "write_plan_id": write_plan["write_plan_id"],
        "write_plan_sha256": _digest(write_plan),
        "source": {
            "filename": filename,
            "sha256": source_sha256,
            "size_bytes": len(source_bytes),
        },
        "output": {
            "filename": output_filename,
            "sha256": output_sha256,
            "size_bytes": len(output_bytes),
        },
        "status": "completed",
        "records": record_receipts,
        "verification": {
            **verification,
            "source_hash_unchanged": hashlib.sha256(source_bytes).hexdigest() == source_sha256,
            "output_reopened": True,
            "external_links_detected": output_map.workbook.external_links_detected,
            "formula_calculation_deferred": True,
        },
        "controls": list(_EXECUTION_CONTROLS),
    }
    receipt = {
        **provisional,
        "execution_id": f"fmre_{_digest(provisional)[:24]}",
    }
    return WorkbookExecutionResult(output_bytes=output_bytes, receipt=receipt)


def execute_workbook_write_plan_file(
    source_path: str | os.PathLike[str],
    *,
    output_path: str | os.PathLike[str],
    write_plan: dict[str, Any],
) -> dict[str, Any]:
    source = Path(source_path)
    output = Path(output_path)
    if source.resolve() == output.resolve():
        raise ValueError("output path must differ from the source path")
    if source.suffix.lower() != ".xlsx" or output.suffix.lower() != ".xlsx":
        raise ValueError("source and output paths must use the .xlsx extension")
    if output.exists():
        raise ValueError("output path already exists")
    if not source.is_file():
        raise ValueError("source workbook does not exist")

    source_bytes = source.read_bytes()
    result = execute_workbook_write_plan_bytes(
        source_bytes,
        filename=source.name,
        output_filename=output.name,
        write_plan=write_plan,
    )

    output.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{output.name}.",
        suffix=".tmp",
        dir=output.parent,
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(result.output_bytes)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, output)
    except Exception:
        temporary.unlink(missing_ok=True)
        output.unlink(missing_ok=True)
        raise
    return result.receipt


def validate_workbook_execution_receipt_payload(
    payload: Any,
    *,
    write_plan: dict[str, Any] | None = None,
) -> tuple[str, ...]:
    issues: list[str] = []
    if not isinstance(payload, dict):
        return ("execution receipt must be an object",)
    allowed = {
        "contract_version",
        "execution_id",
        "write_plan_id",
        "write_plan_sha256",
        "source",
        "output",
        "status",
        "records",
        "verification",
        "controls",
    }
    extras = sorted(set(payload) - allowed)
    if extras:
        issues.append(f"execution receipt contains undeclared fields: {extras}")
    if payload.get("contract_version") != "workbook-execution-receipt.v1":
        issues.append("unsupported contract_version")
    execution_id = payload.get("execution_id")
    if not isinstance(execution_id, str) or not execution_id.startswith("fmre_") or len(execution_id) != 29:
        issues.append("execution_id is invalid")
    if payload.get("status") != "completed":
        issues.append("status must be completed")
    for name in ("source", "output"):
        value = payload.get(name)
        if not isinstance(value, dict):
            issues.append(f"{name} must be an object")
            continue
        if set(value) != {"filename", "sha256", "size_bytes"}:
            issues.append(f"{name} fields are invalid")
        if not isinstance(value.get("filename"), str) or not value.get("filename"):
            issues.append(f"{name}.filename is invalid")
        if not _is_sha256(value.get("sha256")):
            issues.append(f"{name}.sha256 is invalid")
        if not isinstance(value.get("size_bytes"), int) or value.get("size_bytes") <= 0:
            issues.append(f"{name}.size_bytes is invalid")
    records = payload.get("records")
    if not isinstance(records, list):
        issues.append("records must be an array")
    else:
        for index, record in enumerate(records):
            if not isinstance(record, dict):
                issues.append(f"records[{index}] must be an object")
                continue
            expected = {
                "record_id",
                "sequence",
                "write_kind",
                "sheet_name",
                "coordinate",
                "status",
                "cell_count",
                "before_sha256",
                "after_sha256",
            }
            if set(record) != expected:
                issues.append(f"records[{index}] fields are invalid")
            if record.get("status") != "applied":
                issues.append(f"records[{index}].status must be applied")
            for field in ("before_sha256", "after_sha256"):
                if not _is_sha256(record.get(field)):
                    issues.append(f"records[{index}].{field} is invalid")
    verification = payload.get("verification")
    if not isinstance(verification, dict):
        issues.append("verification must be an object")
    else:
        if verification.get("failed_record_ids") != []:
            issues.append("verification.failed_record_ids must be empty")
        for field in (
            "source_hash_unchanged",
            "output_reopened",
            "formula_calculation_deferred",
        ):
            if verification.get(field) is not True:
                issues.append(f"verification.{field} must be true")
        if verification.get("external_links_detected") is not False:
            issues.append("verification.external_links_detected must be false")
    if payload.get("controls") != list(_EXECUTION_CONTROLS):
        issues.append("controls do not match the required control set")
    if write_plan is not None:
        if payload.get("write_plan_id") != write_plan.get("write_plan_id"):
            issues.append("write_plan_id does not match the source write plan")
        if payload.get("write_plan_sha256") != _digest(write_plan):
            issues.append("write_plan_sha256 does not match the source write plan")
        if isinstance(records, list):
            expected_ids = [
                record["record_id"]
                for phase in write_plan.get("phases", [])
                for record in phase.get("records", [])
            ]
            actual_ids = [record.get("record_id") for record in records if isinstance(record, dict)]
            if actual_ids != expected_ids:
                issues.append("receipt record order does not match the write plan")
    if isinstance(execution_id, str) and execution_id.startswith("fmre_"):
        candidate = dict(payload)
        candidate.pop("execution_id", None)
        if execution_id != f"fmre_{_digest(candidate)[:24]}":
            issues.append("execution_id does not match payload")
    return tuple(dict.fromkeys(issues))


def _require_openpyxl() -> None:
    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:
        raise ValueError(
            'Workbook execution dependencies are missing. Install with: pip install -e ".[executor]"'
        ) from exc


def _load_workbook(data: bytes):  # type: ignore[no-untyped-def]
    from openpyxl import load_workbook

    return load_workbook(
        io.BytesIO(data),
        read_only=False,
        data_only=False,
        keep_links=False,
        keep_vba=False,
    )


def _save_workbook(workbook) -> bytes:  # type: ignore[no-untyped-def]
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _request_recalculation(workbook) -> None:  # type: ignore[no-untyped-def]
    calculation = getattr(workbook, "calculation", None)
    if calculation is None:
        return
    for field, value in (
        ("fullCalcOnLoad", True),
        ("forceFullCalc", True),
        ("calcMode", "auto"),
    ):
        if hasattr(calculation, field):
            setattr(calculation, field, value)


def _apply_record(workbook, record: dict[str, Any]) -> None:  # type: ignore[no-untyped-def]
    kind = record["write_kind"]
    sheet_name = record["sheet_name"]
    if kind == "ensure_sheet":
        if sheet_name not in workbook.sheetnames:
            position = record["payload"]["position"]
            index = max(0, min(len(workbook.worksheets), position - 1))
            workbook.create_sheet(title=sheet_name, index=index)
        return
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"write record references a missing sheet: {sheet_name}")
    worksheet = workbook[sheet_name]
    cells = list(_cells(worksheet, record["coordinate"]))
    if kind == "write_value":
        _require_single_cell(cells, record)
        _require_blank_or_same(cells[0], record["payload"]["value"], record)
        cells[0].value = record["payload"]["value"]
    elif kind == "reserve_input":
        for cell in cells:
            if cell.value is not None:
                raise ValueError(f"input reservation is not blank: {record['record_id']}")
    elif kind == "write_formula":
        _require_single_cell(cells, record)
        formula = record["payload"]["formula"]
        _require_blank_or_same(cells[0], formula, record)
        cells[0].value = formula
    elif kind == "apply_style":
        for cell in cells:
            _apply_style(cell, record["payload"]["style"])
    else:
        raise ValueError(f"unsupported write kind: {kind}")


def _apply_style(cell, style: dict[str, Any]) -> None:  # type: ignore[no-untyped-def]
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side

    role = style["role_style"]
    font = role["font"]
    fill = role["fill"]
    alignment = role["alignment"]
    border = role["border"]
    protection = role["protection"]
    bottom_style = border["bottom"]
    cell.font = Font(
        name=font["family"],
        size=font["size"],
        bold=font["bold"],
        italic=font["italic"],
        color=_argb(font["colour"]),
    )
    cell.fill = PatternFill(
        fill_type="solid",
        fgColor=_argb(fill["colour"]),
    )
    cell.alignment = Alignment(
        horizontal=alignment["horizontal"],
        wrap_text=alignment["wrap_text"],
    )
    cell.border = Border(
        bottom=Side(
            style=None if bottom_style == "none" else bottom_style,
            color=_argb("#D1D5DB"),
        )
    )
    cell.protection = Protection(locked=protection["locked"])
    number_format = style["number_format"]["code"]
    if number_format != "source":
        cell.number_format = number_format


def _record_state(workbook, record: dict[str, Any]) -> dict[str, Any]:  # type: ignore[no-untyped-def]
    sheet_name = record["sheet_name"]
    if record["write_kind"] == "ensure_sheet":
        return {
            "sheet_exists": sheet_name in workbook.sheetnames,
            "sheet_position": (
                workbook.sheetnames.index(sheet_name) + 1
                if sheet_name in workbook.sheetnames
                else None
            ),
            "cell_count": 0,
        }
    if sheet_name not in workbook.sheetnames:
        return {"sheet_exists": False, "cell_count": 0, "cells": []}
    worksheet = workbook[sheet_name]
    cells = list(_cells(worksheet, record["coordinate"]))
    return {
        "sheet_exists": True,
        "cell_count": len(cells),
        "cells": [_cell_state(cell) for cell in cells],
    }


def _cell_state(cell) -> dict[str, Any]:  # type: ignore[no-untyped-def]
    return {
        "coordinate": cell.coordinate,
        "value_sha256": _digest(cell.value),
        "data_type": cell.data_type,
        "number_format": cell.number_format,
        "font": {
            "name": cell.font.name,
            "size": cell.font.sz,
            "bold": cell.font.bold,
            "italic": cell.font.italic,
            "colour": _colour_value(cell.font.color),
        },
        "fill": _colour_value(cell.fill.fgColor),
        "alignment": {
            "horizontal": cell.alignment.horizontal,
            "wrap_text": cell.alignment.wrap_text,
        },
        "locked": cell.protection.locked,
    }


def _verify_records(workbook, write_plan: dict[str, Any]) -> dict[str, Any]:  # type: ignore[no-untyped-def]
    verified: list[str] = []
    failed: list[str] = []
    for phase in write_plan["phases"]:
        for record in phase["records"]:
            try:
                _verify_record(workbook, record)
            except ValueError:
                failed.append(record["record_id"])
            else:
                verified.append(record["record_id"])
    return {
        "verified_record_ids": verified,
        "failed_record_ids": failed,
        "verified_record_count": len(verified),
    }


def _verify_record(workbook, record: dict[str, Any]) -> None:  # type: ignore[no-untyped-def]
    sheet_name = record["sheet_name"]
    kind = record["write_kind"]
    if sheet_name not in workbook.sheetnames:
        raise ValueError("sheet missing")
    if kind == "ensure_sheet":
        return
    worksheet = workbook[sheet_name]
    cells = list(_cells(worksheet, record["coordinate"]))
    if kind == "write_value":
        _require_single_cell(cells, record)
        if cells[0].value != record["payload"]["value"]:
            raise ValueError("value mismatch")
    elif kind == "reserve_input":
        if any(cell.value is not None for cell in cells):
            raise ValueError("input reservation not blank")
    elif kind == "write_formula":
        _require_single_cell(cells, record)
        if cells[0].value != record["payload"]["formula"]:
            raise ValueError("formula mismatch")
    elif kind == "apply_style":
        style = record["payload"]["style"]
        for cell in cells:
            if cell.protection.locked != style["role_style"]["protection"]["locked"]:
                raise ValueError("protection mismatch")
            expected_format = style["number_format"]["code"]
            if expected_format != "source" and cell.number_format != expected_format:
                raise ValueError("number format mismatch")


def _cells(worksheet, coordinate: str) -> Iterable[Any]:  # type: ignore[no-untyped-def]
    selected = worksheet[coordinate]
    if isinstance(selected, tuple):
        if selected and isinstance(selected[0], tuple):
            for row in selected:
                yield from row
        else:
            yield from selected
    else:
        yield selected


def _require_single_cell(cells: list[Any], record: dict[str, Any]) -> None:
    if len(cells) != 1:
        raise ValueError(f"write record requires one cell: {record['record_id']}")


def _require_blank_or_same(cell, value: Any, record: dict[str, Any]) -> None:  # type: ignore[no-untyped-def]
    if cell.value not in (None, value):
        raise ValueError(f"write target is not blank: {record['record_id']}")


def _argb(value: str) -> str:
    return "FF" + value.lstrip("#").upper()


def _colour_value(colour) -> str | None:  # type: ignore[no-untyped-def]
    if colour is None:
        return None
    value = getattr(colour, "rgb", None)
    return value if isinstance(value, str) else None


def _is_sha256(value: Any) -> bool:
    return (
        isinstance(value, str)
        and len(value) == 64
        and all(character in "0123456789abcdef" for character in value)
    )


def _digest(payload: Any) -> str:
    rendered = json.dumps(
        payload,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=True,
        default=str,
    ).encode("utf-8")
    return hashlib.sha256(rendered).hexdigest()
