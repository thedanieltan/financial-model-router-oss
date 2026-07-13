from __future__ import annotations

import hashlib
import io
import json
import os
import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from fmr.workbook.executor import _cells
from fmr.workbook.executor_public import validate_workbook_execution_receipt_payload
from fmr.workbook.inspect import inspect_workbook_bytes
from fmr.workbook.write_plan_public import validate_workbook_write_plan_payload

_FORMULA_ERRORS = {
    "#DIV/0!",
    "#N/A",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#REF!",
    "#VALUE!",
}
_NUMERIC_OUTPUT_TYPES = {"currency", "decimal", "multiple", "percentage"}
_ACCEPTANCE_CONTROLS = (
    "calculated_values_not_recorded",
    "calculation_engine_isolated",
    "execution_receipt_pinned",
    "external_links_forbidden",
    "formula_errors_forbidden",
    "immutable_records_verified",
    "input_edits_limited_to_reserved_ranges",
    "output_published_only_after_acceptance",
    "write_plan_pinned",
)


@dataclass(frozen=True)
class CalculationEngine:
    executable: str
    name: str
    version: str
    adapter: str = "libreoffice-headless.v1"

    def to_dict(self) -> dict[str, str]:
        return {
            "name": self.name,
            "version": self.version,
            "adapter": self.adapter,
        }


@dataclass(frozen=True)
class WorkbookCalculationResult:
    output_bytes: bytes
    receipt: dict[str, Any]


def discover_calculation_engine(executable: str | None = None) -> CalculationEngine:
    requested = executable or os.environ.get("FMR_CALCULATION_ENGINE")
    candidates = (requested,) if requested else ("libreoffice", "soffice")
    path = next((shutil.which(item) for item in candidates if item), None)
    if path is None:
        raise ValueError(
            "LibreOffice calculation engine was not found. Install LibreOffice or set FMR_CALCULATION_ENGINE."
        )
    try:
        completed = subprocess.run(
            [path, "--version"],
            check=False,
            capture_output=True,
            text=True,
            timeout=15,
        )
    except (OSError, subprocess.SubprocessError) as exc:
        raise ValueError(f"could not inspect the calculation engine: {exc}") from exc
    version_output = (completed.stdout or completed.stderr).strip().splitlines()
    version = version_output[0].strip() if version_output else "unknown"
    if completed.returncode != 0:
        raise ValueError("calculation engine version check failed")
    return CalculationEngine(
        executable=path,
        name=Path(path).name,
        version=version,
    )


def calculation_engine_status(executable: str | None = None) -> dict[str, Any]:
    try:
        engine = discover_calculation_engine(executable)
    except ValueError as exc:
        return {"available": False, "error": str(exc), "engine": None}
    return {"available": True, "error": None, "engine": engine.to_dict()}


def calculate_and_accept_workbook_bytes(
    input_bytes: bytes,
    *,
    input_filename: str,
    output_filename: str,
    write_plan: dict[str, Any],
    execution_receipt: dict[str, Any],
    engine_executable: str | None = None,
    timeout_seconds: int = 120,
) -> WorkbookCalculationResult:
    engine = discover_calculation_engine(engine_executable)
    output_bytes, process = _run_engine(
        input_bytes,
        input_filename=input_filename,
        engine=engine,
        timeout_seconds=timeout_seconds,
    )
    receipt = accept_calculated_workbook_bytes(
        input_bytes,
        output_bytes,
        input_filename=input_filename,
        output_filename=output_filename,
        write_plan=write_plan,
        execution_receipt=execution_receipt,
        engine={
            **engine.to_dict(),
            "return_code": process["return_code"],
            "stdout_sha256": process["stdout_sha256"],
            "stderr_sha256": process["stderr_sha256"],
        },
    )
    return WorkbookCalculationResult(output_bytes=output_bytes, receipt=receipt)


def calculate_and_accept_workbook_file(
    input_path: str | os.PathLike[str],
    *,
    output_path: str | os.PathLike[str],
    write_plan: dict[str, Any],
    execution_receipt: dict[str, Any],
    engine_executable: str | None = None,
    timeout_seconds: int = 120,
) -> dict[str, Any]:
    source = Path(input_path)
    output = Path(output_path)
    if source.resolve() == output.resolve():
        raise ValueError("calculated output path must differ from the input path")
    if source.suffix.lower() != ".xlsx" or output.suffix.lower() != ".xlsx":
        raise ValueError("input and output paths must use the .xlsx extension")
    if not source.is_file():
        raise ValueError("input workbook does not exist")
    if output.exists():
        raise ValueError("calculated output path already exists")

    result = calculate_and_accept_workbook_bytes(
        source.read_bytes(),
        input_filename=source.name,
        output_filename=output.name,
        write_plan=write_plan,
        execution_receipt=execution_receipt,
        engine_executable=engine_executable,
        timeout_seconds=timeout_seconds,
    )
    if result.receipt["status"] != "passed":
        return result.receipt

    output.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{output.name}.",
        suffix=".tmp",
        dir=output.parent,
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(result.output_bytes)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, output)
    except Exception:
        temporary.unlink(missing_ok=True)
        output.unlink(missing_ok=True)
        raise
    return result.receipt


def accept_calculated_workbook_bytes(
    input_bytes: bytes,
    calculated_bytes: bytes,
    *,
    input_filename: str,
    output_filename: str,
    write_plan: dict[str, Any],
    execution_receipt: dict[str, Any],
    engine: dict[str, Any],
) -> dict[str, Any]:
    if not input_filename.lower().endswith(".xlsx") or not output_filename.lower().endswith(".xlsx"):
        raise ValueError("calculation input and output must use the .xlsx extension")
    if not input_bytes or not calculated_bytes:
        raise ValueError("calculation input and output bytes must be non-empty")

    write_issues = validate_workbook_write_plan_payload(write_plan)
    if write_issues:
        raise ValueError("invalid workbook write plan: " + "; ".join(write_issues))
    receipt_issues = validate_workbook_execution_receipt_payload(
        execution_receipt,
        write_plan=write_plan,
    )
    if receipt_issues:
        raise ValueError("invalid execution receipt: " + "; ".join(receipt_issues))
    _validate_engine_metadata(engine)

    input_map = inspect_workbook_bytes(input_bytes, filename=input_filename)
    output_map = inspect_workbook_bytes(calculated_bytes, filename=output_filename)
    for label, workbook_map in (("input", input_map), ("output", output_map)):
        if workbook_map.external_links_detected:
            raise ValueError(f"{label} workbook contains external links")
        unsupported = [
            item
            for item in workbook_map.findings
            if item.startswith("unsupported_feature:")
            or item.startswith("unsupported_sheet_type:")
        ]
        if unsupported:
            raise ValueError(
                f"{label} workbook contains unsupported features: "
                + "; ".join(unsupported)
            )

    input_formula_workbook = _load_openpyxl(input_bytes, data_only=False)
    output_formula_workbook = _load_openpyxl(calculated_bytes, data_only=False)
    output_value_workbook = _load_openpyxl(calculated_bytes, data_only=True)
    try:
        immutable = _verify_input_records(input_formula_workbook, write_plan)
        formula_checks = _formula_checks(
            output_formula_workbook,
            output_value_workbook,
            write_plan,
        )
        workbook_scan = _scan_all_formula_cells(
            output_formula_workbook,
            output_value_workbook,
        )
    finally:
        input_formula_workbook.close()
        output_formula_workbook.close()
        output_value_workbook.close()

    issue_count = (
        len(immutable["failed_record_ids"])
        + sum(1 for item in formula_checks if item["status"] == "failed")
        + workbook_scan["error_count"]
        + workbook_scan["missing_cached_value_count"]
    )
    status = "passed" if issue_count == 0 else "failed"
    input_sha256 = hashlib.sha256(input_bytes).hexdigest()
    output_sha256 = hashlib.sha256(calculated_bytes).hexdigest()
    provisional = {
        "contract_version": "workbook-calculation-acceptance.v1",
        "write_plan_id": write_plan["write_plan_id"],
        "write_plan_sha256": _digest(write_plan),
        "execution_id": execution_receipt["execution_id"],
        "execution_receipt_sha256": _digest(execution_receipt),
        "input": {
            "filename": input_filename,
            "sha256": input_sha256,
            "size_bytes": len(input_bytes),
            "matches_execution_output_hash": (
                input_sha256 == execution_receipt["output"]["sha256"]
            ),
        },
        "output": {
            "filename": output_filename,
            "sha256": output_sha256,
            "size_bytes": len(calculated_bytes),
        },
        "engine": dict(engine),
        "status": status,
        "summary": {
            "immutable_record_count": immutable["verified_record_count"],
            "immutable_failure_count": len(immutable["failed_record_ids"]),
            "input_cell_count": immutable["input_cell_count"],
            "populated_input_cell_count": immutable["populated_input_cell_count"],
            "planned_formula_count": len(formula_checks),
            "planned_formula_failure_count": sum(
                1 for item in formula_checks if item["status"] == "failed"
            ),
            "workbook_formula_count": workbook_scan["formula_count"],
            "workbook_formula_error_count": workbook_scan["error_count"],
            "missing_cached_value_count": workbook_scan[
                "missing_cached_value_count"
            ],
        },
        "immutable_verification": immutable,
        "formula_checks": formula_checks,
        "workbook_formula_scan": workbook_scan,
        "controls": list(_ACCEPTANCE_CONTROLS),
    }
    return {
        **provisional,
        "acceptance_id": f"fmrc_{_digest(provisional)[:24]}",
    }


def validate_workbook_calculation_acceptance_payload(
    payload: Any,
    *,
    write_plan: dict[str, Any] | None = None,
    execution_receipt: dict[str, Any] | None = None,
) -> tuple[str, ...]:
    issues: list[str] = []
    if not isinstance(payload, dict):
        return ("calculation acceptance must be an object",)
    allowed = {
        "contract_version",
        "acceptance_id",
        "write_plan_id",
        "write_plan_sha256",
        "execution_id",
        "execution_receipt_sha256",
        "input",
        "output",
        "engine",
        "status",
        "summary",
        "immutable_verification",
        "formula_checks",
        "workbook_formula_scan",
        "controls",
    }
    extras = sorted(set(payload) - allowed)
    if extras:
        issues.append(f"calculation acceptance contains undeclared fields: {extras}")
    if payload.get("contract_version") != "workbook-calculation-acceptance.v1":
        issues.append("unsupported contract_version")
    acceptance_id = payload.get("acceptance_id")
    if not isinstance(acceptance_id, str) or not acceptance_id.startswith("fmrc_") or len(acceptance_id) != 29:
        issues.append("acceptance_id is invalid")
    if payload.get("status") not in {"passed", "failed"}:
        issues.append("status is invalid")
    for name in ("input", "output"):
        _validate_file(payload.get(name), name, issues, input_file=name == "input")
    engine = payload.get("engine")
    if not isinstance(engine, dict):
        issues.append("engine must be an object")
    else:
        try:
            _validate_engine_metadata(engine)
        except ValueError as exc:
            issues.append(str(exc))
    summary = payload.get("summary")
    if not isinstance(summary, dict):
        issues.append("summary must be an object")
    else:
        for field in (
            "immutable_record_count",
            "immutable_failure_count",
            "input_cell_count",
            "populated_input_cell_count",
            "planned_formula_count",
            "planned_formula_failure_count",
            "workbook_formula_count",
            "workbook_formula_error_count",
            "missing_cached_value_count",
        ):
            if not isinstance(summary.get(field), int) or summary.get(field) < 0:
                issues.append(f"summary.{field} must be a non-negative integer")
        failed = (
            summary.get("immutable_failure_count", 0)
            + summary.get("planned_formula_failure_count", 0)
            + summary.get("workbook_formula_error_count", 0)
            + summary.get("missing_cached_value_count", 0)
        )
        expected_status = "passed" if failed == 0 else "failed"
        if payload.get("status") != expected_status:
            issues.append("status does not match failure counts")
    formula_checks = payload.get("formula_checks")
    if not isinstance(formula_checks, list):
        issues.append("formula_checks must be an array")
    else:
        for index, check in enumerate(formula_checks):
            _validate_formula_check(check, index, issues)
    if payload.get("controls") != list(_ACCEPTANCE_CONTROLS):
        issues.append("controls do not match the required control set")
    if write_plan is not None:
        if payload.get("write_plan_id") != write_plan.get("write_plan_id"):
            issues.append("write_plan_id does not match the source write plan")
        if payload.get("write_plan_sha256") != _digest(write_plan):
            issues.append("write_plan_sha256 does not match the source write plan")
    if execution_receipt is not None:
        if payload.get("execution_id") != execution_receipt.get("execution_id"):
            issues.append("execution_id does not match the source execution receipt")
        if payload.get("execution_receipt_sha256") != _digest(execution_receipt):
            issues.append("execution_receipt_sha256 does not match the source receipt")
    if isinstance(acceptance_id, str) and acceptance_id.startswith("fmrc_"):
        candidate = dict(payload)
        candidate.pop("acceptance_id", None)
        if acceptance_id != f"fmrc_{_digest(candidate)[:24]}":
            issues.append("acceptance_id does not match payload")
    return tuple(dict.fromkeys(issues))


def _run_engine(
    input_bytes: bytes,
    *,
    input_filename: str,
    engine: CalculationEngine,
    timeout_seconds: int,
) -> tuple[bytes, dict[str, Any]]:
    if timeout_seconds < 1 or timeout_seconds > 600:
        raise ValueError("calculation timeout must be between 1 and 600 seconds")
    with tempfile.TemporaryDirectory(prefix="fmr-calc-") as directory:
        root = Path(directory)
        source_dir = root / "source"
        output_dir = root / "output"
        profile_dir = root / "profile"
        source_dir.mkdir()
        output_dir.mkdir()
        profile_dir.mkdir()
        source_path = source_dir / Path(input_filename).name
        source_path.write_bytes(input_bytes)
        command = [
            engine.executable,
            "--headless",
            "--nologo",
            "--nodefault",
            "--nolockcheck",
            "--nofirststartwizard",
            f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
            "--convert-to",
            "xlsx:Calc MS Excel 2007 XML",
            "--outdir",
            str(output_dir),
            str(source_path),
        ]
        try:
            completed = subprocess.run(
                command,
                check=False,
                capture_output=True,
                timeout=timeout_seconds,
                env={**os.environ, "HOME": str(root)},
            )
        except subprocess.TimeoutExpired as exc:
            raise ValueError("calculation engine timed out") from exc
        except OSError as exc:
            raise ValueError(f"calculation engine could not start: {exc}") from exc
        process = {
            "return_code": completed.returncode,
            "stdout_sha256": hashlib.sha256(completed.stdout).hexdigest(),
            "stderr_sha256": hashlib.sha256(completed.stderr).hexdigest(),
        }
        if completed.returncode != 0:
            raise ValueError(
                f"calculation engine failed with return code {completed.returncode}"
            )
        candidates = sorted(output_dir.glob("*.xlsx"))
        if len(candidates) != 1:
            raise ValueError("calculation engine did not produce exactly one XLSX output")
        output_bytes = candidates[0].read_bytes()
        if not output_bytes:
            raise ValueError("calculation engine produced an empty workbook")
        return output_bytes, process


def _verify_input_records(workbook, write_plan: dict[str, Any]) -> dict[str, Any]:  # type: ignore[no-untyped-def]
    verified: list[str] = []
    failed: list[str] = []
    input_cell_count = 0
    populated_input_cell_count = 0
    for phase in write_plan["phases"]:
        for record in phase["records"]:
            try:
                sheet_name = record["sheet_name"]
                if sheet_name not in workbook.sheetnames:
                    raise ValueError("sheet_missing")
                if record["write_kind"] == "ensure_sheet":
                    verified.append(record["record_id"])
                    continue
                worksheet = workbook[sheet_name]
                cells = list(_cells(worksheet, record["coordinate"]))
                if record["write_kind"] == "reserve_input":
                    input_cell_count += len(cells)
                    populated_input_cell_count += sum(
                        cell.value not in (None, "") for cell in cells
                    )
                elif record["write_kind"] == "write_value":
                    if len(cells) != 1 or cells[0].value != record["payload"]["value"]:
                        raise ValueError("value_mismatch")
                elif record["write_kind"] == "write_formula":
                    if len(cells) != 1 or cells[0].value != record["payload"]["formula"]:
                        raise ValueError("formula_mismatch")
                elif record["write_kind"] == "apply_style":
                    style = record["payload"]["style"]
                    for cell in cells:
                        expected_locked = style["role_style"]["protection"]["locked"]
                        if cell.protection.locked != expected_locked:
                            raise ValueError("protection_mismatch")
                        expected_format = style["number_format"]["code"]
                        if expected_format != "source" and cell.number_format != expected_format:
                            raise ValueError("number_format_mismatch")
                verified.append(record["record_id"])
            except (KeyError, TypeError, ValueError):
                failed.append(record["record_id"])
    if input_cell_count != populated_input_cell_count:
        failed.append("reserved_inputs_incomplete")
    return {
        "verified_record_ids": verified,
        "failed_record_ids": list(dict.fromkeys(failed)),
        "verified_record_count": len(verified),
        "input_cell_count": input_cell_count,
        "populated_input_cell_count": populated_input_cell_count,
    }


def _formula_checks(formula_workbook, value_workbook, write_plan: dict[str, Any]) -> list[dict[str, Any]]:  # type: ignore[no-untyped-def]
    checks: list[dict[str, Any]] = []
    for phase in write_plan["phases"]:
        for record in phase["records"]:
            if record["write_kind"] != "write_formula":
                continue
            sheet_name = record["sheet_name"]
            coordinate = record["coordinate"].split(":", 1)[0]
            issues: list[str] = []
            formula_cell = formula_workbook[sheet_name][coordinate]
            value_cell = value_workbook[sheet_name][coordinate]
            formula_value = formula_cell.value
            calculated_value = value_cell.value
            if not isinstance(formula_value, str) or not formula_value.startswith("="):
                issues.append("formula_missing_after_calculation")
            observed_type = _observed_type(calculated_value)
            if calculated_value is None:
                issues.append("cached_value_missing")
            elif isinstance(calculated_value, str) and calculated_value.upper() in _FORMULA_ERRORS:
                issues.append("formula_error")
            output_type = record["payload"]["output_type"]
            if output_type == "boolean":
                if not isinstance(calculated_value, bool):
                    issues.append("output_type_mismatch")
            elif output_type in _NUMERIC_OUTPUT_TYPES:
                if isinstance(calculated_value, bool) or not isinstance(
                    calculated_value, (int, float)
                ):
                    issues.append("output_type_mismatch")
            sign = _sign(calculated_value)
            if (
                record["payload"]["sign_convention"]
                in {
                    "positive_asset",
                    "positive_expense",
                    "positive_inflow",
                    "positive_liability",
                }
                and sign == "negative"
            ):
                issues.append("sign_convention_mismatch")
            checks.append(
                {
                    "check_id": f"fmrc_{record['record_id']}",
                    "record_id": record["record_id"],
                    "formula_identifier": record["payload"]["formula_identifier"],
                    "sheet_name": sheet_name,
                    "coordinate": record["coordinate"],
                    "status": "passed" if not issues else "failed",
                    "observed_type": observed_type,
                    "observed_sign": sign,
                    "issue_codes": issues,
                }
            )
    return checks


def _scan_all_formula_cells(formula_workbook, value_workbook) -> dict[str, int]:  # type: ignore[no-untyped-def]
    formula_count = 0
    error_count = 0
    missing_count = 0
    for worksheet in formula_workbook.worksheets:
        values = value_workbook[worksheet.title]
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type != "f" and not (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    continue
                formula_count += 1
                value = values[cell.coordinate].value
                if value is None:
                    missing_count += 1
                elif isinstance(value, str) and value.upper() in _FORMULA_ERRORS:
                    error_count += 1
    return {
        "formula_count": formula_count,
        "error_count": error_count,
        "missing_cached_value_count": missing_count,
    }


def _load_openpyxl(data: bytes, *, data_only: bool):  # type: ignore[no-untyped-def]
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            'Calculated-output acceptance requires the executor extra: pip install -e ".[executor]"'
        ) from exc
    return load_workbook(
        io.BytesIO(data),
        read_only=False,
        data_only=data_only,
        keep_links=False,
        keep_vba=False,
    )


def _validate_engine_metadata(engine: Any) -> None:
    if not isinstance(engine, dict):
        raise ValueError("engine must be an object")
    required = {"name", "version", "adapter"}
    if not required.issubset(engine):
        raise ValueError("engine metadata is incomplete")
    allowed = required | {"return_code", "stdout_sha256", "stderr_sha256"}
    extras = sorted(set(engine) - allowed)
    if extras:
        raise ValueError(f"engine metadata contains undeclared fields: {extras}")
    for field in required:
        if not isinstance(engine.get(field), str) or not engine.get(field):
            raise ValueError(f"engine.{field} must be a non-empty string")
    if "return_code" in engine and engine["return_code"] != 0:
        raise ValueError("engine.return_code must be zero")
    for field in ("stdout_sha256", "stderr_sha256"):
        if field in engine and not _is_sha256(engine[field]):
            raise ValueError(f"engine.{field} must be a SHA-256 hex string")


def _validate_file(value: Any, name: str, issues: list[str], *, input_file: bool) -> None:
    if not isinstance(value, dict):
        issues.append(f"{name} must be an object")
        return
    allowed = {"filename", "sha256", "size_bytes"}
    if input_file:
        allowed.add("matches_execution_output_hash")
    extras = sorted(set(value) - allowed)
    if extras:
        issues.append(f"{name} contains undeclared fields: {extras}")
    if not isinstance(value.get("filename"), str) or not value.get("filename"):
        issues.append(f"{name}.filename must be a non-empty string")
    if not _is_sha256(value.get("sha256")):
        issues.append(f"{name}.sha256 must be a SHA-256 hex string")
    if not isinstance(value.get("size_bytes"), int) or value.get("size_bytes") <= 0:
        issues.append(f"{name}.size_bytes must be a positive integer")
    if input_file and not isinstance(value.get("matches_execution_output_hash"), bool):
        issues.append("input.matches_execution_output_hash must be boolean")


def _validate_formula_check(value: Any, index: int, issues: list[str]) -> None:
    context = f"formula_checks[{index}]"
    if not isinstance(value, dict):
        issues.append(f"{context} must be an object")
        return
    expected = {
        "check_id",
        "record_id",
        "formula_identifier",
        "sheet_name",
        "coordinate",
        "status",
        "observed_type",
        "observed_sign",
        "issue_codes",
    }
    if set(value) != expected:
        issues.append(f"{context} fields are invalid")
    if value.get("status") not in {"passed", "failed"}:
        issues.append(f"{context}.status is invalid")
    if not isinstance(value.get("issue_codes"), list) or not all(
        isinstance(item, str) for item in value.get("issue_codes", [])
    ):
        issues.append(f"{context}.issue_codes must be an array of strings")
    elif value.get("status") == "passed" and value["issue_codes"]:
        issues.append(f"{context} passed with issue codes")
    elif value.get("status") == "failed" and not value["issue_codes"]:
        issues.append(f"{context} failed without issue codes")


def _observed_type(value: Any) -> str:
    if value is None:
        return "blank"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, str) and value.upper() in _FORMULA_ERRORS:
        return "error"
    if isinstance(value, str):
        return "text"
    return type(value).__name__.lower()


def _sign(value: Any) -> str:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return "not_applicable"
    if value > 0:
        return "positive"
    if value < 0:
        return "negative"
    return "zero"


def _is_sha256(value: Any) -> bool:
    return (
        isinstance(value, str)
        and len(value) == 64
        and all(character in "0123456789abcdef" for character in value)
    )


def _digest(payload: Any) -> str:
    rendered = json.dumps(
        payload,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=True,
        default=str,
    ).encode("utf-8")
    return hashlib.sha256(rendered).hexdigest()
