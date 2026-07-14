from __future__ import annotations

import csv
import hashlib
import io
import re
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from typing import Any, Iterable

from fmr.data import validate_canonical_financial_data


SOURCE_TYPES = {
    "trial_balance",
    "financial_statement",
    "general_ledger",
    "budget_actual",
    "debt_schedule",
    "operational_driver",
}
SOURCE_SYSTEMS = {"generic", "xero", "quickbooks", "erpnext"}
FORMATS = {"csv", "xlsx"}
FIELDS = {
    "trial_balance": ("period", "account_code", "account_name", "debit", "credit"),
    "financial_statement": ("period", "statement", "concept_id", "value"),
    "general_ledger": ("period", "account_code", "account_name", "amount"),
    "budget_actual": ("period", "concept_id", "actual", "budget"),
    "debt_schedule": (
        "facility_id", "period", "opening_balance", "drawdown", "repayment",
        "interest_rate", "interest", "closing_balance",
    ),
    "operational_driver": ("period", "driver_id", "value"),
}
DECIMAL_FIELDS = {
    "debit", "credit", "value", "amount", "actual", "budget",
    "opening_balance", "drawdown", "repayment", "interest_rate", "interest",
    "closing_balance",
}
IDENTIFIER = re.compile(r"^[a-z0-9]+(?:[._-][a-z0-9]+)*$")


def validate_source_adapter_profile(profile: dict[str, Any]) -> tuple[str, ...]:
    issues: list[str] = []
    expected = {
        "contract_version", "profile_id", "profile_version", "source_system",
        "source_type", "format", "sheet_name", "columns",
    }
    if not isinstance(profile, dict) or profile.get("contract_version") != "source-adapter-profile.v1":
        return ("unsupported source adapter profile contract",)
    if set(profile) != expected:
        issues.append("source adapter profile fields do not match the contract")
    for field in ("profile_id", "profile_version"):
        if not isinstance(profile.get(field), str) or not profile[field]:
            issues.append(f"{field} is required")
    if profile.get("source_system") not in SOURCE_SYSTEMS:
        issues.append("source_system is unsupported")
    source_type = profile.get("source_type")
    if source_type not in SOURCE_TYPES:
        issues.append("source_type is unsupported")
    if profile.get("format") not in FORMATS:
        issues.append("format is unsupported")
    if profile.get("format") == "xlsx" and not isinstance(profile.get("sheet_name"), str):
        issues.append("sheet_name is required for XLSX profiles")
    if profile.get("format") == "csv" and profile.get("sheet_name") is not None:
        issues.append("sheet_name must be null for CSV profiles")
    columns = profile.get("columns")
    required = set(FIELDS.get(source_type, ()))
    if not isinstance(columns, dict) or set(columns) != required:
        issues.append("columns must map every canonical field exactly")
    elif not all(isinstance(value, str) and value.strip() for value in columns.values()):
        issues.append("column headers must be non-empty strings")
    elif len(set(columns.values())) != len(columns):
        issues.append("column headers must be unique")
    return tuple(dict.fromkeys(issues))


def import_tabular_source(
    content: bytes,
    profile: dict[str, Any],
    *,
    source_name: str,
    entity_id: str,
    currency: str,
    entity_name: str | None = None,
) -> dict[str, Any]:
    issues = validate_source_adapter_profile(profile)
    if issues:
        raise ValueError("invalid source adapter profile: " + "; ".join(issues))
    if not source_name or not entity_id or not re.fullmatch(r"[A-Z]{3}", currency):
        raise ValueError("source_name, entity_id and an uppercase three-letter currency are required")
    rows = _read_rows(content, profile)
    if not rows:
        raise ValueError("source export contains no data rows")
    mapped = []
    for number, row in enumerate(rows, start=2):
        record: dict[str, Any] = {}
        for canonical, header in profile["columns"].items():
            value = row.get(header)
            if value is None or (isinstance(value, str) and not value.strip()):
                raise ValueError(f"row {number} field {header!r} is empty")
            record[canonical] = _decimal(value, number, header) if canonical in DECIMAL_FIELDS else str(value).strip()
        mapped.append(record)
    payload = _to_canonical(mapped, profile, entity_id, currency, entity_name)
    payload["provenance"] = [{
        "source": source_name,
        "sha256": hashlib.sha256(content).hexdigest(),
        "source_system": profile["source_system"],
        "profile_id": profile["profile_id"],
        "profile_version": profile["profile_version"],
    }]
    canonical_issues = validate_canonical_financial_data(payload)
    if canonical_issues:
        raise ValueError("canonical financial data is invalid: " + "; ".join(canonical_issues))
    return payload


def merge_canonical_data(
    packages: Iterable[dict[str, Any]], *, assumptions: dict[str, Any] | None = None
) -> dict[str, Any]:
    values = list(packages)
    if not values:
        raise ValueError("at least one canonical package is required")
    for package in values:
        issues = validate_canonical_financial_data(package)
        if issues:
            raise ValueError("invalid canonical package: " + "; ".join(issues))
    entity = values[0]["entity"]
    if any(item["entity"] != entity for item in values[1:]):
        raise ValueError("canonical packages must have identical entity metadata")
    periods = list(dict.fromkeys(period for item in values for period in item["periods"]))
    statements: dict[str, dict[str, dict[str, Any]]] = defaultdict(lambda: defaultdict(dict))
    drivers: dict[str, dict[str, Any]] = defaultdict(dict)
    for package in values:
        for statement, concepts in package["financial_statements"].items():
            for concept, series in concepts.items():
                _merge_series(statements[statement][concept], package["periods"], series, f"{statement}.{concept}")
        for driver, series in package["operational_drivers"].items():
            _merge_series(drivers[driver], package["periods"], series, f"operational_drivers.{driver}")
    merged = _empty_canonical(entity["entity_id"], entity["currency"], entity.get("entity_name"), periods)
    merged["financial_statements"] = {
        statement: {concept: _complete_series(by_period, periods, f"{statement}.{concept}") for concept, by_period in concepts.items()}
        for statement, concepts in statements.items()
    }
    merged["operational_drivers"] = {name: _complete_series(by_period, periods, f"operational_drivers.{name}") for name, by_period in drivers.items()}
    for section in ("trial_balance", "account_balances", "debt_schedules", "capital_expenditure", "working_capital"):
        merged[section] = [record for package in values for record in package[section]]
    merged["assumptions"] = dict(assumptions or {})
    merged["provenance"] = [record for package in values for record in package["provenance"]]
    issues = validate_canonical_financial_data(merged)
    if issues:
        raise ValueError("merged canonical data is invalid: " + "; ".join(issues))
    return merged


def _read_rows(content: bytes, profile: dict[str, Any]) -> list[dict[str, Any]]:
    if profile["format"] == "csv":
        try:
            reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
        except UnicodeDecodeError as exc:
            raise ValueError("CSV source must be UTF-8 encoded") from exc
        headers = reader.fieldnames or []
        if len(headers) != len(set(headers)):
            raise ValueError("source export contains duplicate headers")
        _require_headers(headers, profile)
        return [dict(row) for row in reader]
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("XLSX ingestion requires the openpyxl optional dependency") from exc
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
    if profile["sheet_name"] not in workbook.sheetnames:
        raise ValueError(f"XLSX sheet {profile['sheet_name']!r} is missing")
    sheet = workbook[profile["sheet_name"]]
    iterator = sheet.iter_rows(values_only=False)
    try:
        header_cells = next(iterator)
    except StopIteration as exc:
        raise ValueError("XLSX source is empty") from exc
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_cells]
    if len(headers) != len(set(headers)):
        raise ValueError("source export contains duplicate headers")
    _require_headers(headers, profile)
    result = []
    for row_number, cells in enumerate(iterator, start=2):
        record = {}
        for index, header in enumerate(headers):
            cell = cells[index]
            if cell.data_type == "f" and header in profile["columns"].values():
                raise ValueError(f"row {row_number} mapped field {header!r} contains an unevaluated formula")
            record[header] = cell.value
        if any(value is not None and value != "" for value in record.values()):
            result.append(record)
    return result


def _require_headers(headers: list[str], profile: dict[str, Any]) -> None:
    missing = sorted(set(profile["columns"].values()) - set(headers))
    if missing:
        raise ValueError("source export is missing mapped headers: " + ", ".join(missing))


def _decimal(value: Any, row: int, header: str) -> str:
    if isinstance(value, bool):
        raise ValueError(f"row {row} field {header!r} is not a decimal")
    try:
        parsed = Decimal(str(value).strip().replace(",", ""))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"row {row} field {header!r} is not a decimal") from exc
    if not parsed.is_finite():
        raise ValueError(f"row {row} field {header!r} is not finite")
    return format(parsed, "f")


def _to_canonical(rows: list[dict[str, Any]], profile: dict[str, Any], entity_id: str, currency: str, entity_name: str | None) -> dict[str, Any]:
    source_type = profile["source_type"]
    periods = list(dict.fromkeys(row["period"] for row in rows))
    payload = _empty_canonical(entity_id, currency, entity_name, periods)
    if source_type == "financial_statement":
        indexed: dict[tuple[str, str], dict[str, str]] = defaultdict(dict)
        for row in rows:
            if row["statement"] not in {"income_statement", "balance_sheet", "cash_flow"}:
                raise ValueError(f"unsupported statement {row['statement']!r}")
            if not IDENTIFIER.fullmatch(row["concept_id"]):
                raise ValueError(f"invalid concept_id {row['concept_id']!r}")
            _put_unique(indexed[(row["statement"], row["concept_id"])], row["period"], row["value"], "statement series")
        payload["financial_statements"] = {
            statement: {concept: _complete_series(by_period, periods, concept) for (kind, concept), by_period in indexed.items() if kind == statement}
            for statement in ("income_statement", "balance_sheet", "cash_flow")
            if any(kind == statement for kind, _ in indexed)
        }
    elif source_type == "operational_driver":
        indexed = defaultdict(dict)
        for row in rows:
            if not IDENTIFIER.fullmatch(row["driver_id"]):
                raise ValueError(f"invalid driver_id {row['driver_id']!r}")
            _put_unique(indexed[row["driver_id"]], row["period"], row["value"], "driver series")
        payload["operational_drivers"] = {name: _complete_series(values, periods, name) for name, values in indexed.items()}
    elif source_type == "trial_balance":
        balances = defaultdict(lambda: [Decimal("0"), Decimal("0")])
        for row in rows:
            if Decimal(row["debit"]) < 0 or Decimal(row["credit"]) < 0:
                raise ValueError("trial balance debit and credit must be non-negative")
            balances[row["period"]][0] += Decimal(row["debit"])
            balances[row["period"]][1] += Decimal(row["credit"])
        unbalanced = [period for period, (debit, credit) in balances.items() if debit != credit]
        if unbalanced:
            raise ValueError("trial balance is unbalanced for periods: " + ", ".join(unbalanced))
        payload["trial_balance"] = rows
    elif source_type == "general_ledger":
        payload["account_balances"] = rows
    elif source_type == "budget_actual":
        payload["working_capital"] = rows
    else:
        seen: set[tuple[str, str]] = set()
        for row in rows:
            key = (row["facility_id"], row["period"])
            if key in seen:
                raise ValueError("debt schedule contains duplicate facility and period")
            seen.add(key)
            expected = Decimal(row["opening_balance"]) + Decimal(row["drawdown"]) - Decimal(row["repayment"])
            if expected != Decimal(row["closing_balance"]):
                raise ValueError(f"debt schedule roll-forward fails for {key[0]} in {key[1]}")
        payload["debt_schedules"] = rows
    return payload


def _empty_canonical(entity_id: str, currency: str, entity_name: str | None, periods: list[str]) -> dict[str, Any]:
    entity = {"entity_id": entity_id, "currency": currency}
    if entity_name is not None:
        entity["entity_name"] = entity_name
    return {
        "contract_version": "canonical-financial-data.v2", "entity": entity,
        "periods": periods, "financial_statements": {}, "trial_balance": [],
        "account_balances": [], "debt_schedules": [], "capital_expenditure": [],
        "working_capital": [], "operational_drivers": {}, "assumptions": {},
        "provenance": [{"source": "pending"}],
    }


def _put_unique(target: dict[str, Any], period: str, value: Any, label: str) -> None:
    if period in target:
        raise ValueError(f"{label} contains duplicate period {period!r}")
    target[period] = value


def _complete_series(values: dict[str, Any], periods: list[str], label: str) -> list[Any]:
    missing = [period for period in periods if period not in values]
    if missing:
        raise ValueError(f"{label} is missing periods: " + ", ".join(missing))
    return [values[period] for period in periods]


def _merge_series(target: dict[str, Any], periods: list[str], values: list[Any], label: str) -> None:
    for period, value in zip(periods, values, strict=True):
        if period in target and Decimal(str(target[period])) != Decimal(str(value)):
            raise ValueError(f"conflicting value for {label} in {period}")
        target[period] = value
