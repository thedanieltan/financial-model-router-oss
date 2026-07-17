from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any


@dataclass(frozen=True)
class SaasBudgetInputs:
    company_name: str
    currency: str = "SGD"
    forecast_months: int = 12
    opening_arr: Decimal = Decimal("1200000")
    monthly_new_arr: Decimal = Decimal("50000")
    monthly_expansion_arr: Decimal = Decimal("15000")
    monthly_contraction_arr: Decimal = Decimal("5000")
    monthly_churned_arr: Decimal = Decimal("10000")
    gross_margin_rate: Decimal = Decimal("0.80")
    sales_marketing_monthly: Decimal = Decimal("80000")
    rnd_monthly: Decimal = Decimal("90000")
    ga_monthly: Decimal = Decimal("50000")
    starting_cash: Decimal = Decimal("1500000")
    headcount: Decimal = Decimal("25")
    average_salary_per_head_monthly: Decimal = Decimal("9000")


def _decimal(value: Any, field: str) -> Decimal:
    try:
        result = Decimal(str(value))
    except (InvalidOperation, TypeError) as exc:
        raise ValueError(f"{field} must be numeric") from exc
    if not result.is_finite():
        raise ValueError(f"{field} must be finite")
    return result


def _rate(value: Any, field: str) -> Decimal:
    result = _decimal(value, field)
    if result < Decimal("0") or result > Decimal("1"):
        raise ValueError(f"{field} must be between 0 and 1")
    return result


def _money(value: Decimal) -> float:
    return float(value.quantize(Decimal("0.01")))


def _input(payload: dict[str, Any], name: str, default: Decimal) -> Decimal:
    return _decimal(payload.get(name, default), name)


def parse_saas_budget_inputs(payload: dict[str, Any]) -> SaasBudgetInputs:
    company_name = str(payload.get("company_name") or "Example SaaS Company").strip()
    if not company_name:
        raise ValueError("company_name is required")
    currency = str(payload.get("currency") or "SGD").strip().upper()
    if not currency or len(currency) > 8:
        raise ValueError("currency must be a short currency code or label")
    months_raw = payload.get("forecast_months", 12)
    if isinstance(months_raw, bool) or not isinstance(months_raw, int):
        raise ValueError("forecast_months must be an integer")
    if months_raw < 1 or months_raw > 60:
        raise ValueError("forecast_months must be between 1 and 60")
    return SaasBudgetInputs(
        company_name=company_name,
        currency=currency,
        forecast_months=months_raw,
        opening_arr=_input(payload, "opening_arr", Decimal("1200000")),
        monthly_new_arr=_input(payload, "monthly_new_arr", Decimal("50000")),
        monthly_expansion_arr=_input(payload, "monthly_expansion_arr", Decimal("15000")),
        monthly_contraction_arr=_input(payload, "monthly_contraction_arr", Decimal("5000")),
        monthly_churned_arr=_input(payload, "monthly_churned_arr", Decimal("10000")),
        gross_margin_rate=_rate(payload.get("gross_margin_rate", Decimal("0.80")), "gross_margin_rate"),
        sales_marketing_monthly=_input(payload, "sales_marketing_monthly", Decimal("80000")),
        rnd_monthly=_input(payload, "rnd_monthly", Decimal("90000")),
        ga_monthly=_input(payload, "ga_monthly", Decimal("50000")),
        starting_cash=_input(payload, "starting_cash", Decimal("1500000")),
        headcount=_input(payload, "headcount", Decimal("25")),
        average_salary_per_head_monthly=_input(payload, "average_salary_per_head_monthly", Decimal("9000")),
    )


def _assert_openpyxl() -> tuple[Any, Any, Any, Any, Any, Any]:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError('SaaS workbook generation requires: pip install -e ".[executor]"') from exc
    return Workbook, load_workbook, Alignment, Font, PatternFill, get_column_letter


def build_saas_budget_workbook_bytes(inputs: SaasBudgetInputs) -> bytes:
    Workbook, load_workbook, Alignment, Font, PatternFill, get_column_letter = _assert_openpyxl()
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    assumptions = wb.create_sheet("Assumptions")
    arr = wb.create_sheet("ARR Bridge")
    revenue = wb.create_sheet("Revenue Forecast")
    opex = wb.create_sheet("Opex & Headcount")
    cash = wb.create_sheet("Cash Runway")
    scenarios = wb.create_sheet("Scenarios")
    checks = wb.create_sheet("Checks")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    bold = Font(bold=True)

    def style_header(ws: Any, row: int = 1) -> None:
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    assumptions.append(["Assumption", "Value", "Notes"])
    assumption_rows = [
        ("Company name", inputs.company_name, "Display only"),
        ("Currency", inputs.currency, "Display only"),
        ("Forecast months", inputs.forecast_months, "1 to 60"),
        ("Opening ARR", _money(inputs.opening_arr), "ARR at start of forecast"),
        ("Monthly new ARR", _money(inputs.monthly_new_arr), "Base case"),
        ("Monthly expansion ARR", _money(inputs.monthly_expansion_arr), "Base case"),
        ("Monthly contraction ARR", _money(inputs.monthly_contraction_arr), "Base case"),
        ("Monthly churned ARR", _money(inputs.monthly_churned_arr), "Base case"),
        ("Gross margin rate", float(inputs.gross_margin_rate), "0.0 to 1.0"),
        ("Sales & marketing monthly", _money(inputs.sales_marketing_monthly), "Base case"),
        ("R&D monthly", _money(inputs.rnd_monthly), "Base case"),
        ("G&A monthly", _money(inputs.ga_monthly), "Base case"),
        ("Starting cash", _money(inputs.starting_cash), "Cash at start"),
        ("Headcount", float(inputs.headcount), "Opening headcount"),
        ("Average salary per head monthly", _money(inputs.average_salary_per_head_monthly), "Fully loaded"),
    ]
    for row in assumption_rows:
        assumptions.append(list(row))
    style_header(assumptions)
    assumptions[1][0].font = bold

    month_headers = [f"Month {i}" for i in range(1, inputs.forecast_months + 1)]
    arr.append(["Metric", *month_headers])
    arr_rows = [
        "Opening ARR",
        "New ARR",
        "Expansion ARR",
        "Contraction ARR",
        "Churned ARR",
        "Ending ARR",
        "Gross revenue retention",
        "Net revenue retention",
        "ARR growth vs opening",
    ]
    for metric in arr_rows:
        arr.append([metric])
    style_header(arr)
    for idx in range(2, inputs.forecast_months + 2):
        col = get_column_letter(idx)
        prior = "Assumptions!$B$4" if idx == 2 else f"{get_column_letter(idx - 1)}7"
        arr[f"{col}2"] = f"={prior}"
        arr[f"{col}3"] = "=Assumptions!$B$5"
        arr[f"{col}4"] = "=Assumptions!$B$6"
        arr[f"{col}5"] = "=Assumptions!$B$7"
        arr[f"{col}6"] = "=Assumptions!$B$8"
        arr[f"{col}7"] = f"={col}2+{col}3+{col}4-{col}5-{col}6"
        arr[f"{col}8"] = f"=IF({col}2=0,0,({col}2-{col}5-{col}6)/{col}2)"
        arr[f"{col}9"] = f"=IF({col}2=0,0,({col}2+{col}3+{col}4-{col}5-{col}6)/{col}2)"
        arr[f"{col}10"] = f"=IF(Assumptions!$B$4=0,0,{col}7/Assumptions!$B$4-1)"

    revenue.append(["Metric", *month_headers])
    for metric in ["Average ARR", "Revenue", "Gross margin rate", "Gross profit"]:
        revenue.append([metric])
    style_header(revenue)
    for idx in range(2, inputs.forecast_months + 2):
        col = get_column_letter(idx)
        revenue[f"{col}2"] = f"=('ARR Bridge'!{col}2+'ARR Bridge'!{col}7)/2"
        revenue[f"{col}3"] = f"={col}2/12"
        revenue[f"{col}4"] = "=Assumptions!$B$9"
        revenue[f"{col}5"] = f"={col}3*{col}4"

    opex.append(["Metric", *month_headers])
    for metric in ["Headcount", "Payroll", "Sales & marketing", "R&D", "G&A", "Total opex", "Sales efficiency"]:
        opex.append([metric])
    style_header(opex)
    for idx in range(2, inputs.forecast_months + 2):
        col = get_column_letter(idx)
        opex[f"{col}2"] = "=Assumptions!$B$14"
        opex[f"{col}3"] = f"={col}2*Assumptions!$B$15"
        opex[f"{col}4"] = "=Assumptions!$B$10"
        opex[f"{col}5"] = "=Assumptions!$B$11"
        opex[f"{col}6"] = "=Assumptions!$B$12"
        opex[f"{col}7"] = f"=SUM({col}3:{col}6)"
        opex[f"{col}8"] = f"=IF({col}4=0,0,'ARR Bridge'!{col}3/{col}4)"

    cash.append(["Metric", *month_headers])
    for metric in ["Opening cash", "Gross profit", "Total opex", "EBITDA proxy", "Free cash flow proxy", "Ending cash", "Runway months"]:
        cash.append([metric])
    style_header(cash)
    for idx in range(2, inputs.forecast_months + 2):
        col = get_column_letter(idx)
        prior_cash = "Assumptions!$B$13" if idx == 2 else f"{get_column_letter(idx - 1)}7"
        cash[f"{col}2"] = f"={prior_cash}"
        cash[f"{col}3"] = f"='Revenue Forecast'!{col}5"
        cash[f"{col}4"] = f"='Opex & Headcount'!{col}7"
        cash[f"{col}5"] = f"={col}3-{col}4"
        cash[f"{col}6"] = f"={col}5"
        cash[f"{col}7"] = f"={col}2+{col}6"
        cash[f"{col}8"] = f"=IF({col}6>=0,999,{col}7/ABS({col}6))"

    scenarios.append(["Scenario", "New ARR multiplier", "Churn multiplier", "Ending ARR", "Ending cash"])
    scenario_rows = [("Downside", 0.75, 1.25), ("Base", 1.00, 1.00), ("Upside", 1.25, 0.80)]
    last_col = get_column_letter(inputs.forecast_months + 1)
    for name, new_mult, churn_mult in scenario_rows:
        scenarios.append([name, new_mult, churn_mult, None, None])
    style_header(scenarios)
    # Keep scenarios simple and reviewable: link base case outputs and expose multipliers for sensitivity extension.
    for row in range(2, 5):
        scenarios[f"D{row}"] = f"='ARR Bridge'!{last_col}7*B{row}/C{row}"
        scenarios[f"E{row}"] = f"='Cash Runway'!{last_col}7"

    checks.append(["Check", "Status", "Review note"])
    checks.append(["ARR bridge ties", f"=AND('ARR Bridge'!B7=B2+B3+B4-B5-B6)", "Ending ARR must equal opening + new + expansion - contraction - churn"])
    checks.append(["Gross margin in range", "=AND(Assumptions!B9>=0,Assumptions!B9<=1)", "Gross margin must be 0% to 100%"])
    checks.append(["Cash runway computed", f"=ISNUMBER('Cash Runway'!{last_col}8)", "Runway should be numeric"])
    checks.append(["No negative ending ARR", f"=MIN('ARR Bridge'!B7:{last_col}7)>=0", "Negative ARR needs review"])
    checks.append(["Scenario output changes", "=AND(Scenarios!D2<>Scenarios!D3,Scenarios!D4<>Scenarios!D3)", "Upside/downside should move from base"])
    style_header(checks)

    summary.append([f"{inputs.company_name} SaaS Budget & Forecast"])
    summary["A1"].font = title_font
    summary.append(["Currency", "=Assumptions!B2"])
    summary.append(["Forecast months", "=Assumptions!B3"])
    summary.append(["Opening ARR", "=Assumptions!B4"])
    summary.append(["Ending ARR", f"='ARR Bridge'!{last_col}7"])
    summary.append(["ARR growth", f"='ARR Bridge'!{last_col}10"])
    summary.append(["Latest NRR", f"='ARR Bridge'!{last_col}9"])
    summary.append(["Latest GRR", f"='ARR Bridge'!{last_col}8"])
    summary.append(["Ending cash", f"='Cash Runway'!{last_col}7"])
    summary.append(["Runway months", f"='Cash Runway'!{last_col}8"])
    summary.append(["Management memo"])
    summary.append(["This workbook is a deterministic planning artefact. Review the assumptions, checks and formulas before relying on outputs."])
    summary["A11"].font = bold

    percent_rows = {"ARR Bridge": [8, 9, 10], "Revenue Forecast": [4], "Assumptions": [9]}
    money_sheets = ["ARR Bridge", "Revenue Forecast", "Opex & Headcount", "Cash Runway", "Scenarios", "Summary", "Assumptions"]
    for ws in wb.worksheets:
        ws.freeze_panes = "B2"
        for column in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(column)].width = 18 if column > 1 else 30
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=1).font = bold if row > 1 else ws.cell(row=row, column=1).font
        if ws.title in money_sheets:
            for row in ws.iter_rows(min_row=2):
                for cell in row[1:]:
                    if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                        cell.number_format = '#,##0.00'
        for row_idx in percent_rows.get(ws.title, []):
            for cell in ws[row_idx][1:]:
                cell.number_format = '0.0%'
    for cell in scenarios[1]:
        cell.font = header_font
        cell.fill = header_fill

    output = BytesIO()
    wb.save(output)
    data = output.getvalue()
    wb.close()
    validate_saas_budget_workbook_bytes(data)
    return data


def validate_saas_budget_workbook_bytes(data: bytes) -> dict[str, Any]:
    _, load_workbook, _, _, _, _ = _assert_openpyxl()
    wb = load_workbook(BytesIO(data), data_only=False, read_only=True)
    try:
        required = {
            "Summary",
            "Assumptions",
            "ARR Bridge",
            "Revenue Forecast",
            "Opex & Headcount",
            "Cash Runway",
            "Scenarios",
            "Checks",
        }
        issues: list[str] = []
        missing = sorted(required - set(wb.sheetnames))
        if missing:
            issues.append("missing_sheets:" + ",".join(missing))
        formulas = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                formulas += sum(isinstance(cell.value, str) and cell.value.startswith("=") for cell in row)
        if formulas < 30:
            issues.append("insufficient_formula_coverage")
        if not missing:
            checks = wb["Checks"]
            expected_checks = {
                "ARR bridge ties",
                "Gross margin in range",
                "Cash runway computed",
                "No negative ending ARR",
                "Scenario output changes",
            }
            actual_checks = {str(checks.cell(row=row, column=1).value) for row in range(2, checks.max_row + 1)}
            if not expected_checks <= actual_checks:
                issues.append("missing_required_checks")
            arr = wb["ARR Bridge"]
            if arr["B7"].value != "=B2+B3+B4-B5-B6":
                issues.append("formula_mismatch:ARR Bridge!B7")
            cash = wb["Cash Runway"]
            if cash["B8"].value != "=IF(B6>=0,999,B7/ABS(B6))":
                issues.append("formula_mismatch:Cash Runway!B8")
        return {"valid": not issues, "issues": issues, "sheet_count": len(wb.sheetnames), "formula_count": formulas}
    finally:
        wb.close()


def build_saas_budget_workbook_from_payload(payload: dict[str, Any]) -> bytes:
    return build_saas_budget_workbook_bytes(parse_saas_budget_inputs(payload))
