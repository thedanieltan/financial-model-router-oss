from __future__ import annotations

import hashlib
import json
import os
import tempfile
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from fmr.data import validate_canonical_model_input


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _load_input(reference: dict[str, Any]) -> dict[str, Any]:
    path = reference.get("path")
    expected_hash = reference.get("sha256")
    if not isinstance(path, str) or not path or not isinstance(expected_hash, str) or len(expected_hash) != 64:
        raise ValueError("canonical model input reference requires path and sha256")
    data = Path(path).read_bytes()
    if _sha256(data) != expected_hash:
        raise ValueError("canonical model input hash mismatch")
    payload = json.loads(data)
    issues = validate_canonical_model_input(payload)
    if issues:
        raise ValueError("invalid canonical model input: " + "; ".join(issues))
    return payload


def _decimal(value: Any, field: str) -> Decimal:
    try:
        result = Decimal(str(value))
    except (InvalidOperation, TypeError) as exc:
        raise ValueError(f"{field} must be a decimal") from exc
    if not result.is_finite():
        raise ValueError(f"{field} must be finite")
    return result


def _forecast_inputs(model_input: dict[str, Any]) -> dict[str, Any]:
    assumptions = model_input["assumptions"]
    horizon = assumptions.get("forecast_horizon")
    if isinstance(horizon, bool) or not isinstance(horizon, int) or horizon < 1:
        raise ValueError("forecast_horizon must be a positive integer")
    scenario = assumptions.get("scenario")
    if scenario not in {"base", "upside", "downside"}:
        raise ValueError("scenario must be base, upside or downside")
    adjustments = assumptions.get("scenario_adjustments")
    if not isinstance(adjustments, dict) or not isinstance(adjustments.get(scenario), dict):
        raise ValueError("scenario_adjustments must explicitly define the selected scenario")
    selected = adjustments[scenario]
    statement = model_input["financial_statements"].get("income_statement", {})
    if not isinstance(statement.get("revenue"), list) or not isinstance(statement.get("operating_costs"), list):
        raise ValueError("revenue and operating_costs history are required")
    revenue_growth = _decimal(assumptions.get("revenue_growth_rate"), "revenue_growth_rate")
    cost_growth = _decimal(assumptions.get("operating_cost_growth_rate"), "operating_cost_growth_rate")
    revenue_delta = _decimal(selected.get("revenue_growth_delta"), "revenue_growth_delta")
    cost_delta = _decimal(selected.get("operating_cost_growth_delta"), "operating_cost_growth_delta")
    periods = _forecast_periods(model_input["periods"][-1], horizon)
    return {
        "horizon": horizon, "scenario": scenario, "periods": periods,
        "revenue_growth": revenue_growth, "cost_growth": cost_growth,
        "revenue_delta": revenue_delta, "cost_delta": cost_delta,
        "last_revenue": _decimal(statement["revenue"][-1], "historical revenue"),
        "last_cost": _decimal(statement["operating_costs"][-1], "historical operating costs"),
    }


def _forecast_values(inputs: dict[str, Any]) -> list[dict[str, str]]:
    revenue = inputs["last_revenue"]
    cost = inputs["last_cost"]
    rows = []
    for period in inputs["periods"]:
        revenue *= Decimal(1) + inputs["revenue_growth"] + inputs["revenue_delta"]
        cost *= Decimal(1) + inputs["cost_growth"] + inputs["cost_delta"]
        rows.append({
            "period": period,
            "revenue": str(revenue.quantize(Decimal("0.01"))),
            "operating_costs": str(cost.quantize(Decimal("0.01"))),
            "operating_profit": str((revenue - cost).quantize(Decimal("0.01"))),
        })
    return rows


def validate_budget_workbook(path: str | os.PathLike[str]) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError('Native XLSX requires the "executor" package extra') from exc
    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        required = {"Historical Actuals", "Assumptions", "Budget Forecast", "Checks"}
        missing = sorted(required - set(workbook.sheetnames))
        issues: list[str] = []
        formulas = 0
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                formulas += sum(isinstance(cell.value, str) and cell.value.startswith("=") for cell in row)
        if missing:
            issues.append("missing_sheets:" + ",".join(missing))
        if formulas == 0:
            issues.append("no_formulas")
        if not missing:
            forecast = workbook["Budget Forecast"]
            count = max(forecast.max_column - 1, 0)
            if count < 1 or any(forecast.cell(2, column).value != "Forecast" for column in range(2, count + 2)):
                issues.append("actual_forecast_separation_missing")
            for index, column in enumerate(_columns(count)):
                prior_revenue = f"'Historical Actuals'!{_columns(workbook['Historical Actuals'].max_column - 1)[-1]}2" if index == 0 else f"{_columns(count)[index - 1]}3"
                prior_cost = f"'Historical Actuals'!{_columns(workbook['Historical Actuals'].max_column - 1)[-1]}3" if index == 0 else f"{_columns(count)[index - 1]}4"
                expected = {
                    3: f"={prior_revenue}*(1+Assumptions!$B$2+Assumptions!$B$5)",
                    4: f"={prior_cost}*(1+Assumptions!$B$3+Assumptions!$B$6)",
                    5: f"={column}3-{column}4",
                }
                for row, formula in expected.items():
                    if forecast[f"{column}{row}"].value != formula:
                        issues.append(f"formula_mismatch:Budget Forecast!{column}{row}")
            if workbook["Checks"]["B2"].value != f"=COLUMNS('Budget Forecast'!B1:{_columns(count)[-1]}1)=Assumptions!B4":
                issues.append("formula_mismatch:Checks!B2")
        return {"status": "passed" if not issues else "failed", "issues": issues, "sheet_count": len(workbook.sheetnames), "formula_count": formulas}
    finally:
        workbook.close()


def execute_budget_forecast_handoff(handoff: dict[str, Any], output_dir: str | os.PathLike[str]) -> dict[str, Any]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise RuntimeError('Native XLSX requires the "executor" package extra') from exc
    payload = handoff.get("provider_payload")
    if not isinstance(payload, dict) or payload.get("adapter_id") != "native-xlsx/generic-budget-forecast.v2":
        raise ValueError("unsupported Native XLSX adapter")
    model_input = _load_input(payload.get("canonical_financial_data", {}))
    inputs = _forecast_inputs(model_input)
    forecast_values = _forecast_values(inputs)
    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)
    basename = payload.get("output_basename", "budget-forecast")
    workbook_output = destination / f"{basename}.xlsx"
    json_output = destination / f"{basename}.json"
    if workbook_output.exists() or json_output.exists():
        raise ValueError("output path already exists")

    workbook = Workbook()
    actuals = workbook.active
    actuals.title = "Historical Actuals"
    assumptions = workbook.create_sheet("Assumptions")
    forecast = workbook.create_sheet("Budget Forecast")
    checks = workbook.create_sheet("Checks")
    historical_periods = model_input["periods"]
    statement = model_input["financial_statements"]["income_statement"]
    actuals.append(["Metric", *historical_periods])
    actuals.append(["Revenue", *[float(value) for value in statement["revenue"]]])
    actuals.append(["Operating costs", *[float(value) for value in statement["operating_costs"]]])
    assumptions.append(["Assumption", "Value"])
    assumptions.append(["Revenue growth rate", float(inputs["revenue_growth"])])
    assumptions.append(["Operating cost growth rate", float(inputs["cost_growth"])])
    assumptions.append(["Forecast horizon", inputs["horizon"]])
    assumptions.append(["Scenario revenue delta", float(inputs["revenue_delta"])])
    assumptions.append(["Scenario operating cost delta", float(inputs["cost_delta"])])
    assumptions.append(["Scenario", inputs["scenario"]])
    forecast.append(["Metric", *inputs["periods"]])
    forecast.append(["Period type", *(["Forecast"] * inputs["horizon"])])
    columns = _columns(inputs["horizon"])
    historical_last = _columns(len(historical_periods))[-1]
    revenue_sources = [f"'Historical Actuals'!{historical_last}2", *[f"{column}3" for column in columns[:-1]]]
    cost_sources = [f"'Historical Actuals'!{historical_last}3", *[f"{column}4" for column in columns[:-1]]]
    forecast.append(["Revenue"] + [f"={source}*(1+Assumptions!$B$2+Assumptions!$B$5)" for source in revenue_sources])
    forecast.append(["Operating costs"] + [f"={source}*(1+Assumptions!$B$3+Assumptions!$B$6)" for source in cost_sources])
    forecast.append(["Operating profit"] + [f"={column}3-{column}4" for column in columns])
    checks.append(["Check", "Status"])
    checks.append(["Forecast horizon", f"=COLUMNS('Budget Forecast'!B1:{columns[-1]}1)=Assumptions!B4"])
    checks.append(["All periods forecast", f'=COUNTIF(\'Budget Forecast\'!B2:{columns[-1]}2,"Forecast")=Assumptions!B4'])
    for sheet in (actuals, assumptions, forecast, checks):
        for cell in sheet[1]:
            cell.font = Font(bold=True)

    with tempfile.NamedTemporaryFile(prefix=".fmr-", suffix=".xlsx", dir=destination, delete=False) as handle:
        temporary = Path(handle.name)
    try:
        workbook.save(temporary)
        workbook.close()
        validation = validate_budget_workbook(temporary)
        if validation["status"] != "passed":
            raise ValueError("Native XLSX output validation failed: " + "; ".join(validation["issues"]))
        workbook_data = temporary.read_bytes()
        os.replace(temporary, workbook_output)
    finally:
        workbook.close()
        temporary.unlink(missing_ok=True)
    json_document = {
        "contract_version": "budget-forecast-result.v1", "scenario": inputs["scenario"],
        "actual_periods": historical_periods, "forecast_periods": inputs["periods"], "forecast": forecast_values,
    }
    json_data = (json.dumps(json_document, indent=2, sort_keys=True) + "\n").encode()
    with tempfile.NamedTemporaryFile(prefix=".fmr-", suffix=".json", dir=destination, delete=False) as handle:
        json_temporary = Path(handle.name)
        handle.write(json_data)
    try:
        os.replace(json_temporary, json_output)
    finally:
        json_temporary.unlink(missing_ok=True)
    return {
        "provider_receipt_version": "native-xlsx-receipt.v2", "status": "completed", "handoff_sha256": handoff["handoff_sha256"],
        "output_artifacts": [
            {"kind": "budget_forecast_workbook", "format": "xlsx", "path": str(workbook_output), "sha256": _sha256(workbook_data), "size_bytes": len(workbook_data)},
            {"kind": "budget_forecast", "format": "json", "path": str(json_output), "sha256": _sha256(json_data), "size_bytes": len(json_data)},
        ],
        "validation": {**validation, "checks": ["actual_forecast_separation", "forecast_reconciliation", "scenario_applied"]},
        "controls": ["atomic_output", "input_hash_verified", "no_input_values_in_receipt", "output_reopened_and_validated"],
    }


def _forecast_periods(last: str, count: int) -> list[str]:
    if last.isdigit() and len(last) == 4:
        return [str(int(last) + offset) for offset in range(1, count + 1)]
    return [f"Forecast {offset}" for offset in range(1, count + 1)]


def _columns(count: int) -> list[str]:
    result = []
    for number in range(2, count + 2):
        letters = ""
        value = number
        while value:
            value, remainder = divmod(value - 1, 26)
            letters = chr(65 + remainder) + letters
        result.append(letters)
    return result
