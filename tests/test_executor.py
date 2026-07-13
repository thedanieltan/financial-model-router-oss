from __future__ import annotations

import hashlib
import json
import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook

from fmr.types import ModelRequest
from fmr.workbook import (
    analyse_workbook_map,
    compile_workbook_patch,
    compile_workbook_write_plan,
    execute_workbook_write_plan_bytes,
    execute_workbook_write_plan_file,
    inspect_workbook_bytes,
    plan_workbook_content,
    plan_workbook_coordinates,
    plan_workbook_realization,
    resolve_workbook_patch_targets,
    validate_workbook_execution_receipt_payload,
)
from tests.xlsx_factory import financial_workbook


def executable_workbook() -> bytes:
    workbook = Workbook()
    income = workbook.active
    income.title = "Income Statement"
    income["A1"] = "Income Statement"
    income.merge_cells("A1:D1")
    income["B2"] = 2024
    income["C2"] = 2025
    income["D2"] = "2026E"
    income["A3"] = "Revenue"
    income["B3"] = 100
    income["C3"] = 120
    income["D3"] = "=C3*1.10"
    income["A4"] = "EBITDA"
    income["B4"] = 20
    income["C4"] = 24
    income["D4"] = "=D3*0.20"
    income["A5"] = "Net Income"

    balance = workbook.create_sheet("Balance Sheet")
    balance.sheet_state = "hidden"
    balance["A1"] = "Balance Sheet"
    balance["B2"] = 2024
    balance["C2"] = 2025
    balance["A3"] = "Total Assets"
    balance["A4"] = "Total Liabilities"
    balance["A5"] = "Total Equity"

    assumptions = workbook.create_sheet("Assumptions")
    assumptions["A1"] = "Assumptions"
    assumptions["A3"] = "Growth Rate"
    assumptions["A4"] = "Tax Rate"
    assumptions["A5"] = "WACC"

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _write_plan_for(source_bytes: bytes) -> dict:
    workbook_map = inspect_workbook_bytes(source_bytes, filename="synthetic.xlsx")
    request = ModelRequest(
        objective="build a budget forecast",
        role="finance_manager",
        available_data=(
            "balance_sheet_history",
            "revenue_drivers",
            "operating_cost_drivers",
        ),
        workbook_capabilities=(),
        assumptions=("forecast_horizon",),
    )
    analysis = analyse_workbook_map(workbook_map, request)
    patch = compile_workbook_patch(analysis).to_dict()
    resolution = resolve_workbook_patch_targets(analysis, patch).to_dict()
    coordinates = plan_workbook_coordinates(
        analysis,
        patch,
        resolution,
        forecast_period_count=5,
    )
    content = plan_workbook_content(coordinates)
    realization = plan_workbook_realization(content)
    bindings: dict[str, dict] = {}
    for operation in realization["operation_realizations"]:
        for slot in operation["slots"]:
            formula = slot.get("formula_binding")
            if not isinstance(formula, dict):
                continue
            for dependency in formula["dependencies"]:
                if dependency["binding_type"] in {"content_slot", "period_context"}:
                    continue
                bindings[dependency["identifier"]] = {
                    "binding_type": "constant",
                    "value": True if dependency["binding_type"] == "validation_context" else 1,
                }
    context = {
        "contract_version": "workbook-write-context.v1",
        "period_labels": [f"P{index}" for index in range(1, 13)],
        "bindings": bindings,
    }
    return compile_workbook_write_plan(realization, context)


def execution_case() -> tuple[bytes, dict]:
    source_bytes = executable_workbook()
    return source_bytes, _write_plan_for(source_bytes)


class WorkbookExecutorTests(unittest.TestCase):
    def test_executor_writes_copy_and_emits_valid_receipt(self) -> None:
        source_bytes, write_plan = execution_case()
        source_hash = hashlib.sha256(source_bytes).hexdigest()
        result = execute_workbook_write_plan_bytes(
            source_bytes,
            filename="synthetic.xlsx",
            output_filename="completed.xlsx",
            write_plan=write_plan,
        )

        self.assertEqual(hashlib.sha256(source_bytes).hexdigest(), source_hash)
        self.assertNotEqual(result.output_bytes, source_bytes)
        self.assertEqual(
            validate_workbook_execution_receipt_payload(
                result.receipt,
                write_plan=write_plan,
            ),
            (),
        )
        self.assertEqual(result.receipt["status"], "completed")
        self.assertEqual(result.receipt["verification"]["failed_record_ids"], [])
        self.assertEqual(
            result.receipt["verification"]["verified_record_count"],
            write_plan["write_record_count"],
        )

        workbook = load_workbook(BytesIO(result.output_bytes), data_only=False)
        try:
            formula_records = [
                record
                for phase in write_plan["phases"]
                for record in phase["records"]
                if record["write_kind"] == "write_formula"
            ]
            self.assertTrue(formula_records)
            first = formula_records[0]
            cell = workbook[first["sheet_name"]][first["coordinate"].split(":", 1)[0]]
            self.assertEqual(cell.value, first["payload"]["formula"])
        finally:
            workbook.close()

    def test_file_executor_never_overwrites_source_or_existing_output(self) -> None:
        source_bytes, write_plan = execution_case()
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "source.xlsx"
            output = root / "output.xlsx"
            source.write_bytes(source_bytes)
            source_hash = hashlib.sha256(source.read_bytes()).hexdigest()

            receipt = execute_workbook_write_plan_file(
                source,
                output_path=output,
                write_plan=write_plan,
            )
            self.assertTrue(output.is_file())
            self.assertEqual(hashlib.sha256(source.read_bytes()).hexdigest(), source_hash)
            self.assertEqual(receipt["output"]["sha256"], hashlib.sha256(output.read_bytes()).hexdigest())

            with self.assertRaisesRegex(ValueError, "already exists"):
                execute_workbook_write_plan_file(
                    source,
                    output_path=output,
                    write_plan=write_plan,
                )
            with self.assertRaisesRegex(ValueError, "must differ"):
                execute_workbook_write_plan_file(
                    source,
                    output_path=source,
                    write_plan=write_plan,
                )

    def test_source_hash_mismatch_fails_without_output(self) -> None:
        source_bytes, write_plan = execution_case()
        altered = source_bytes + b"x"
        with self.assertRaisesRegex(ValueError, "hash does not match"):
            execute_workbook_write_plan_bytes(
                altered,
                filename="synthetic.xlsx",
                output_filename="completed.xlsx",
                write_plan=write_plan,
            )

    def test_detected_chart_is_rejected(self) -> None:
        source_bytes = financial_workbook(include_chart=True)
        write_plan = _write_plan_for(source_bytes)
        with self.assertRaisesRegex(ValueError, "unsupported_feature:charts"):
            execute_workbook_write_plan_bytes(
                source_bytes,
                filename="synthetic.xlsx",
                output_filename="completed.xlsx",
                write_plan=write_plan,
            )

    def test_receipt_contains_hashes_not_cell_values(self) -> None:
        source_bytes, write_plan = execution_case()
        result = execute_workbook_write_plan_bytes(
            source_bytes,
            filename="synthetic.xlsx",
            output_filename="completed.xlsx",
            write_plan=write_plan,
        )
        rendered = json.dumps(result.receipt, sort_keys=True)
        self.assertNotIn("before_value", rendered)
        self.assertNotIn("after_value", rendered)
        for record in result.receipt["records"]:
            self.assertEqual(
                set(record),
                {
                    "record_id",
                    "sequence",
                    "write_kind",
                    "sheet_name",
                    "coordinate",
                    "status",
                    "cell_count",
                    "before_sha256",
                    "after_sha256",
                },
            )


if __name__ == "__main__":
    unittest.main()
