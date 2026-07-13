from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from fmr.workbook import execute_workbook_write_plan_bytes
from tests.test_executor import execution_case


def populated_execution_case(*, input_value: int = 7) -> tuple[bytes, dict, dict]:
    source_bytes, write_plan = execution_case()
    execution = execute_workbook_write_plan_bytes(
        source_bytes,
        filename="synthetic.xlsx",
        output_filename="executed.xlsx",
        write_plan=write_plan,
    )
    workbook = load_workbook(BytesIO(execution.output_bytes), data_only=False)
    try:
        for phase in write_plan["phases"]:
            for record in phase["records"]:
                if record["write_kind"] != "reserve_input":
                    continue
                worksheet = workbook[record["sheet_name"]]
                min_column, min_row, max_column, max_row = range_boundaries(
                    record["coordinate"]
                )
                for row in worksheet.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_column,
                    max_col=max_column,
                ):
                    for cell in row:
                        cell.value = input_value
        stream = BytesIO()
        workbook.save(stream)
    finally:
        workbook.close()
    return stream.getvalue(), write_plan, execution.receipt


def tamper_first_generated_value(workbook_bytes: bytes, write_plan: dict) -> bytes:
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    try:
        target = next(
            record
            for phase in write_plan["phases"]
            for record in phase["records"]
            if record["write_kind"] == "write_value"
        )
        coordinate = target["coordinate"].split(":", 1)[0]
        workbook[target["sheet_name"]][coordinate] = "tampered"
        stream = BytesIO()
        workbook.save(stream)
    finally:
        workbook.close()
    return stream.getvalue()
